"""
excel_reader.py
---------------
قراءة ملف Excel للحصول على عناوين الفيديوهات بالترتيب.
يحفظ آخر رقم مُستخدم في data/excel_pointer.json
عند الوصول للنهاية، يعود للصف الأول تلقائياً.
"""
import json
import logging
from pathlib import Path
from openpyxl import load_workbook

log = logging.getLogger("excel_reader")

ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
EXCEL_PATH = DATA_DIR / "videos.xlsx"
POINTER_PATH = DATA_DIR / "excel_pointer.json"


def ensure_excel_exists():
    """التحقق من وجود ملف Excel."""
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(
            f"❌ Excel file not found: {EXCEL_PATH}\n"
            f"Please upload videos.xlsx to data/ folder with columns:\n"
            f"  Column A: رقم (number)\n"
            f"  Column B: العنوان (title)\n"
            f"  Column C: جملة بديلة (read text)"
        )


def load_pointer() -> int:
    """تحميل آخر رقم تم استخدامه."""
    try:
        if POINTER_PATH.exists():
            data = json.loads(POINTER_PATH.read_text(encoding="utf-8"))
            return int(data.get("last_used", 0))
    except Exception as e:
        log.warning(f"Failed to load pointer: {e}, starting from 0")
    return 0


def save_pointer(last_used: int):
    """حفظ آخر رقم تم استخدامه."""
    POINTER_PATH.parent.mkdir(parents=True, exist_ok=True)
    POINTER_PATH.write_text(
        json.dumps({"last_used": last_used}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def read_all_rows() -> list[dict]:
    """
    قراءة كل الصفوف من ملف Excel.
    يتجاهل الصف الأول (Header).
    
    Returns:
        list of dicts: [{"number": 1, "title": "...", "read_text": "..."}, ...]
    """
    ensure_excel_exists()
    
    workbook = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    sheet = workbook.active
    
    rows = []
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        # تجاهل الصف الأول (Header)
        if idx == 1:
            continue
        
        # تجاهل الصفوف الفارغة
        if not row or all(cell is None for cell in row):
            continue
        
        # استخراج الأعمدة الثلاثة
        try:
            number = row[0]
            title = row[1]
            read_text = row[2] if len(row) > 2 else None
            
            # تحقق من القيم الأساسية
            if number is None or title is None:
                log.warning(f"Row {idx}: missing data, skipping")
                continue
            
            rows.append({
                "number": int(number),
                "title": str(title).strip(),
                "read_text": str(read_text).strip() if read_text else "اقرأ الوصف",
            })
        except (ValueError, TypeError) as e:
            log.warning(f"Row {idx}: invalid data ({e}), skipping")
            continue
    
    workbook.close()
    
    if not rows:
        raise RuntimeError(f"❌ Excel file is empty or has no valid rows: {EXCEL_PATH}")
    
    log.info(f"✓ Loaded {len(rows)} rows from {EXCEL_PATH.name}")
    return rows


def get_next_video() -> dict:
    """
    جلب الفيديو التالي بالترتيب.
    عند الوصول للنهاية، يعود للصف الأول.
    
    Returns:
        dict: {"number": int, "title": str, "read_text": str}
    """
    rows = read_all_rows()
    total = len(rows)
    
    last_used = load_pointer()
    log.info(f"📊 Excel: {total} rows | Last used: #{last_used}")
    
    # حساب الرقم التالي
    # إذا تجاوزنا النهاية، نبدأ من الأول
    next_index = last_used % total  # 0-based index
    
    selected = rows[next_index]
    new_last_used = last_used + 1
    
    # حفظ المؤشر الجديد
    save_pointer(new_last_used)
    
    log.info(
        f"✓ Selected video #{selected['number']}: {selected['title'][:50]}... "
        f"(iteration {new_last_used})"
    )
    
    return selected


def reset_pointer():
    """إعادة تعيين المؤشر للبداية (للاختبار)."""
    save_pointer(0)
    log.info("✓ Pointer reset to 0")


# =========================
# Test
# =========================
if __name__ == "__main__":
    import sys
    
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
        handlers=[logging.StreamHandler(sys.stdout)],
    )
    
    print("=" * 60)
    print("📊 Excel Reader Test")
    print("=" * 60)
    
    try:
        # اقرأ كل الصفوف
        rows = read_all_rows()
        print(f"\n✓ Total rows: {len(rows)}\n")
        
        for row in rows[:5]:  # أول 5 صفوف
            print(f"  #{row['number']}: {row['title']}")
            print(f"     read: {row['read_text']}\n")
        
        if len(rows) > 5:
            print(f"  ... and {len(rows) - 5} more")
        
        # اختبار get_next_video (5 مرات لرؤية الدوران)
        print("\n" + "=" * 60)
        print("🎲 Testing get_next_video (5 calls):")
        print("=" * 60)
        
        for i in range(1, 6):
            print(f"\n--- Call {i} ---")
            video = get_next_video()
            print(f"  Number: {video['number']}")
            print(f"  Title:  {video['title']}")
            print(f"  Read:   {video['read_text']}")
        
        print("\n✅ Test completed!")
        
    except Exception as e:
        print(f"\n❌ Error: {e}")
        sys.exit(1)
